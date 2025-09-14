###     library imported
import json, re, collections, audioop, pyaudio, vosk, pypinyin, sys, pyttsx3, keyboard, time
from openpyxl.utils import get_column_letter, column_index_from_string
from Levenshtein import distance
from openpyxl import load_workbook
import pandas as pd

SETTINGS = r"Resources\settings.txt"
NUMBERS = r"Resources\numbers.txt"
NAMES = r"Resources\names.txt"
MODEL = r"Resources\vosk-model-cn-0.22"
ENCODING = "utf-8"
FOLDER = "Excel"            #可能更改优化
TIMES = 1
Settings = ["FOLDER", ' ', "file_path", ' ',
            "sheet_name", ' ', "NAME_COL", ' ',
            "NAME_COL_START", ' ', "NAME_COL_END", ' ',]


file_path = FOLDER + "\\"
sheet_name = "Sheet1"

###     function definition
def refresh_name_txt():
    START_POINT = NAME_COL + str(NAME_COL_START)
    start_row, start_col = addr_to_indices(START_POINT)
    END_POINT = NAME_COL + str(NAME_COL_END)
    end_row, end_col = addr_to_indices(END_POINT)
    nrows = 1 + end_row - start_row
    df_write = pd.read_excel(file_path, sheet_name=sheet_name,
                             skiprows=start_row - 1, usecols=NAME_COL,
                             nrows=nrows)
    list_names = []
    list_names.append(df_write.columns[0])
    for i in range(nrows - 1):
        list_names.append(df_write.iloc[i, 0])
    # print(list_names)
    name2txt(list_names)

def reset():
    f = open(SETTINGS, 'w', encoding=ENCODING)
    f.write(str(FOLDER));f.write("\n")
    f.write(str(file_path));f.write("\n")
    f.write(str(sheet_name));f.write("\n")
    f.write(str(NAME_COL));f.write("\n")
    f.write(str(NAME_COL_START));f.write("\n")
    f.write(str(NAME_COL_END));f.write("\n")
    f.close()

def readsettings():
    with open(SETTINGS, 'r', encoding=ENCODING) as f:
        setting_list = [l.strip() for l in f if l.strip()]
    return setting_list

def addr_to_indices(cell_addr):
    """'B3' → (3, 2)  (row, col)"""
    from openpyxl.utils import range_boundaries
    c, r = range_boundaries(cell_addr + ':' + cell_addr)[:2]
    return r, c

def sound2num(str) -> int:   #str具有空格 点有可能被误判成零，请注意
    num = 0
    key = 0
    i = 1
    for w in str.split():
        if w == ' ':
            continue
        else:
            if key == 1:
                if w in NUM_LIST[:10]:
                    num += NUM_LIST.index(w) * 0.1 ** i
                    i += 1
                else:
                    print("""something went wrong 
                          错误代码01
                          小数点之后的数似乎不应该这么读""")
            elif w in NUM_LIST[:10]:
                num += NUM_LIST.index(w)
            elif w == NUM_LIST[10]:
                num *= 10
            elif w == NUM_LIST[11]:
                num *= 100
            elif w == NUM_LIST[12]:
                key = 1
            else:
                pass
    return num
#test
#print(sound2num("一 百 零 五 点 五 六 七"))

# print(voice2name())
def voice2name(namelist) -> str:
    voice = []
    #print("请说出要查询的名字……")
    while True:
        data = stream_name.read(400, exception_on_overflow=False)
        ring.append(audioop.rms(data, 2))
        if sum(ring) / len(ring) < 500:
            if voice:
                rec.AcceptWaveform(b''.join(voice))
                q = json.loads(rec.FinalResult()).get("text", "")
                q = re.sub(r'\s+', '', q)
                if q: break
                voice.clear()
        else:
            voice.append(data)
    # return names   模糊匹配
    if not namelist:
        raise RuntimeError("姓名名单为空，请检查 names.txt 是否正确加载！")
    qpy = ''.join(pypinyin.lazy_pinyin(q))
    return min(namelist, key=lambda n: 0.7 * distance(qpy, PY_INDEX[n]) + 0.3 * distance(q, n))

# print(sound2num(voice2sound()))
def voice2sound() -> str:
    #print("请连续说出由关键字组成的句子：")
    while True:
        data = stream_num.read(800, exception_on_overflow=False)
        if rec.AcceptWaveform(data):
            sent = json.loads(rec.Result())["text"].replace(" ", "")
            if sent:
                break

    # 3. 逐字比对（允许 1 个拼音编辑距离误差）
    THRESH = 1
    result = []
    for ch in sent:
        if ch in keys:  # 完全匹配
            result.append(ch)
            continue
        # 拼音容错
        ch_py = ''.join(pypinyin.lazy_pinyin(ch))
        best, best_d = None, 999
        for k, k_py in key_py.items():
            d = distance(ch_py, k_py)
            if d < best_d:
                best, best_d = k, d
        if best_d <= THRESH:
            result.append(best)
        # 否则丢弃或记录异常
    #print("原始文本：", sent)
    #print("逐字关键字：", " ".join(result))
    return " ".join(result)

def findfiles(string) -> list:
    from pathlib import Path
    folder = Path(string)  # 改成你的目录
    xlsx_files = list(folder.glob('*.xlsx'))  # 注意大小写
    xlsx_file_names = []
    for i in xlsx_files:
        str = i.name
        if str[:2] == "~$":
            pass
        else:
            xlsx_file_names.append(str)
    #print('共找到 xlsx:', len(xlsx_files))
    #for i in xlsx_files:
    #    print(i.name)
    return xlsx_file_names
#print(findfiles("Excel"))

def findSheets(string) -> list:
    with pd.ExcelFile(string) as xls:
        all_sheets = xls.sheet_names
    return all_sheets
#print(findSheets(r"Excel\00.xlsx"))

def name2txt(names: list) -> None:
    f = open(NAMES,'w',encoding=ENCODING)
    #print(names)
    for name in names:
        f.write(str(name))
        f.write("\n")
    f.close()


def speak(text: str,
          rate: int = 200,      # 语速（词/分钟）
          volume: float = 0.9   # 音量 0.0~1.0
         ) -> None:
    if not text.strip():
        return
    engine = pyttsx3.init()
    engine.setProperty("rate", rate)
    engine.setProperty("volume", volume)

    # 可选：列出并选择声音
    # voices = engine.getProperty("voices")
    # for i, v in enumerate(voices):
    #     print(i, v.name, v.languages)
    # engine.setProperty("voice", voices[0].id)

    engine.say(text)
    engine.runAndWait()


def speak00(audio):
    en = pyttsx3.init()
    en.say(audio)
    en.runAndWait()
    en.stop()
    del en

def key():
    while True:                             ##为什么返回同一个值
        event = keyboard.read_event(suppress= False)
        if event.event_type == "down":
            if event.name == 'space':
                return "read"
            elif event.name == 's':
                return "interrupt"              #退到上一步
            elif event.name == 'w':
                return "write"
        #elif keyboard.is_pressed('r'):
            #return "renew"

def loopforrecord(NAME_LIST):
    #print("您希望成绩录入哪一列：")
    #score_col = input()
    #name_list = NAME_LIST.copy()
    score_backup = ""
    record = []

    book = load_workbook(file_path)
    file = book[sheet_name]

    while len(name_list) > 0:
        key_value = key()
        #print(key_value)
        if key_value == "read":
            name = voice2name(name_list)
            print(name)
            speak(name)
        elif key_value == "interrupt":
            time.sleep(TIMES)
            if len(record) ==0 :
                print("已经撤回此次操作")
                continue
            record.pop()
            name = record.pop()
            name_list.append(name)
            print("已经撤回了", name, "的成绩，请重新开始")
            continue
        elif key_value == "write":
            time.sleep(TIMES)
            print("请输入姓名：",end="")
            name = input()
            name = re.sub(r'[A-Za-z0-9\s]+', '', name)
            if name not in name_list:
                print("你输入的名字似乎已经录过成绩或不存在，请重新读名字")
                print(NAME_LIST, name_list, name)
                continue
            print("你输入的名字是", name, "请检查是否正确")
            print("若错误请按 s 撤销操作，若正确请继续")
        else:
            time.sleep(TIMES)
            print("\n你似乎按错了按键，请重新读名字")
            continue

        key_value = key()
        if key_value == "interrupt":
            time.sleep(TIMES)
            print("请重新读姓名")
            continue
        elif key_value == "read":
            SCORE = voice2sound()
            sound = []
            for ch in SCORE.split():
                if ch != ' ':
                    sound.append(ch)
            print(SCORE)
            score = sound2num(SCORE)
            speak(str(sound))
        elif key_value == "write":
            time.sleep(TIMES)
            try:
                print("请输入成绩：",end="")
                score = input()
                score = float( re.sub(r'[A-Za-z0-9\s]+', '', score))
                print("你输入的成绩是", score, "请检查是否正确")
                print("若错误请按 s 撤销操作")
            except ValueError:
                score = 0
                print("请按 s 撤销操作，若正确请继续")
        else:
            print("你似乎按错了按键，请重新读名字")
            continue

        #准备下一步
        record.append(name)
        record.append(' ')
        name_list.remove(name)
        score_backup = score

        #输入excel表格
        file[score_col + str(NAME_LIST.index(name) + start_row)] = score
        book.save(file_path)
        print("录入成功",name, score)




#settings
print("请按照菜单对应的数字敲击数字以进入设置：")
print("1 -> 暂时修改配置")
print("2 -> 更改默认设置")
print("如果您不需要修改任何配置，请您输入期望的成绩存放地址：")
keys = input()

if keys == "1" or keys == "2":
    #选中文件
    xlsx_file_names = findfiles(FOLDER)
    xlsx_num = len(xlsx_file_names)
    if xlsx_num == 0:
        print("文件夹内部未检测到.xlsx文件，请检查文件命名等问题！")
        sys.exit()
    elif xlsx_num > 1:
        print("扫描到多个文件，请输入前置数字以打开对应文件：")
        i = 1
        for file in xlsx_file_names:
            print(i," -> ",file)
            i += 1
        i = int(input())
        file_path += xlsx_file_names[i-1]
        print("已成功选中文件：", file_path)
    elif xlsx_num == 1:
        file_path += xlsx_file_names[0]
        print("已自动扫描并选中文件：", file_path)
    else:
        pass
    #寻找Sheet
    sheet_names = findSheets(file_path)
    sheet_num = len(sheet_names)
    if sheet_num > 1:
        print("扫描到多个表格，请输入前置数字以打开对应表格：")
        i = 1
        for sheet in sheet_names:
            print(i, " -> ", sheet)
            i += 1
        i = int(input())
        sheet_name = sheet_names[i-1]
        print("已成功选中表格：", sheet_name)
    elif sheet_num == 1:
        sheet_name = sheet_names[0]
        print("已自动扫描并选中文件：", sheet_name)
    else:
        print("读取Sheets表格出现故障")
        sys.exit()
    #加载姓名
    print("请输入学生姓名所在的列（如A,D,F）：")
    NAME_COL = input()
    print("请输入学生所在列的起点：")
    NAME_COL_START = int(input())
    print("请输入学生所在列的终点：")
    NAME_COL_END = int(input())

    if keys == "2":
        reset()

    refresh_name_txt()
    print("设置修改完毕，请输入您期望的成绩存储位置：")
    score_col = input()

else:
    setting_list = readsettings()
    FOLDER = setting_list[0]
    file_path = setting_list[1]
    sheet_name = setting_list[2]
    NAME_COL = setting_list[3]
    NAME_COL_START = int(setting_list[4])
    NAME_COL_END = int(setting_list[5])
    score_col = keys

print("主程序启动")

START_POINT = NAME_COL + str(NAME_COL_START)
start_row, start_col = addr_to_indices(START_POINT)
END_POINT = NAME_COL + str(NAME_COL_END)
end_row, end_col = addr_to_indices(END_POINT)
nrows = 1 + end_row - start_row
df_write = pd.read_excel(file_path, sheet_name=sheet_name,
                        skiprows=start_row - 1, usecols=NAME_COL,
                        nrows=nrows)
list_names = []
list_names.append(df_write.columns[0])
for i in range(nrows - 1):
    list_names.append(df_write.iloc[i, 0])
# print(list_names)
name2txt(list_names)
print("姓名加载完成")

























###     initializing
###     加载数字索引、姓名、关键字
with open(NUMBERS, encoding=ENCODING) as f:
    NUM_LIST = [l.strip() for l in f if l.strip()]
with open(NAMES, encoding=ENCODING) as f:               ###其实多此一举，但是我懒得改了
    NAME_LIST = [l.strip() for l in f if l.strip()]
PY_INDEX = {n: ''.join(pypinyin.lazy_pinyin(n)) for n in NAME_LIST}
with open(NUMBERS, encoding=ENCODING) as f:
    keys = [line.strip() for line in f if line.strip()]
key_py = {k: ''.join(pypinyin.lazy_pinyin(k)) for k in keys}

###     加载语音模型
model = vosk.Model(MODEL)
rec = vosk.KaldiRecognizer(model, 16000)
p = pyaudio.PyAudio()
stream_name = p.open(format=pyaudio.paInt16, channels=1, rate=16000, input=True, frames_per_buffer=8000)
stream_num = p.open(format=pyaudio.paInt16, channels=1,rate=16000, input=True, frames_per_buffer=800)
ring = collections.deque(maxlen=8)
#加载音响
#engine = pyttsx3.init()
#engine.setProperty('voice', 'zh')       # 尝试中文
#test
#print(voice2name())
#print(sound2num(voice2sound()))
print("语音模型加载完成")

###     工作区
print("总人数：", len(NAME_LIST), "您可以检查姓名：")
print(NAME_LIST)
print("资源加载完毕，轻敲空格以开始录入成绩")

#print(voice2name(NAME_LIST))
name_list = NAME_LIST.copy()
loopforrecord(NAME_LIST)

print("本次成绩录入还剩", len(name_list), "个人未录取")
print("他们分别是：", name_list)

###     运行结束
stream_name.stop_stream(); stream_name.close()
stream_num.stop_stream(); stream_num.close()
p.terminate()
print("程序运行结束，祝您生活愉快！")